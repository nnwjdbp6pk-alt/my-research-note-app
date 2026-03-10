from __future__ import annotations

import csv
import io
from typing import Any, Literal

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from .db import get_db
from . import schemas, crud

router = APIRouter(prefix="/api")


# --- Project API ---

@router.get("/projects", response_model=list[schemas.ProjectOut], summary="전체 프로젝트 목록 조회")
def list_projects(db: Session = Depends(get_db)):
    return crud.list_projects(db)


@router.post("/projects", response_model=schemas.ProjectOut, status_code=status.HTTP_201_CREATED, summary="프로젝트 생성")
def create_project(payload: schemas.ProjectCreate, db: Session = Depends(get_db)):
    return crud.create_project(db, payload)


@router.get("/projects/{project_id}", response_model=schemas.ProjectOut, summary="프로젝트 상세 조회")
def get_project(project_id: int, db: Session = Depends(get_db)):
    obj = crud.get_project(db, project_id)
    if not obj:
        raise HTTPException(status_code=404, detail="해당 프로젝트를 찾을 수 없습니다.")
    return obj


@router.patch("/projects/{project_id}", response_model=schemas.ProjectOut, summary="프로젝트 수정")
def update_project(project_id: int, payload: schemas.ProjectUpdate, db: Session = Depends(get_db)):
    obj = crud.update_project(db, project_id, payload)
    if not obj:
        raise HTTPException(status_code=404, detail="수정할 프로젝트를 찾을 수 없습니다.")
    return obj


@router.delete("/projects/{project_id}", summary="프로젝트 삭제")
def delete_project(project_id: int, db: Session = Depends(get_db)):
    ok = crud.delete_project(db, project_id)
    if not ok:
        raise HTTPException(status_code=404, detail="삭제할 프로젝트를 찾을 수 없습니다.")
    return {"ok": True, "message": "프로젝트가 삭제되었습니다."}


# --- Experiment API ---

@router.get("/projects/{project_id}/experiments", response_model=list[schemas.ExperimentOut], summary="프로젝트별 실험 목록 조회")
def list_experiments(project_id: int, db: Session = Depends(get_db)):
    return crud.list_experiments(db, project_id)


@router.post("/experiments", response_model=schemas.ExperimentOut, status_code=status.HTTP_201_CREATED, summary="실험 생성")
def create_experiment(payload: schemas.ExperimentCreate, db: Session = Depends(get_db)):
    proj = crud.get_project(db, payload.project_id)
    if not proj:
        raise HTTPException(status_code=400, detail="유효하지 않은 프로젝트 ID입니다.")
    try:
        return crud.create_experiment(db, payload)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.get("/experiments/{experiment_id}", response_model=schemas.ExperimentOut, summary="실험 상세 조회")
def get_experiment(experiment_id: int, db: Session = Depends(get_db)):
    obj = crud.get_experiment(db, experiment_id)
    if not obj:
        raise HTTPException(status_code=404, detail="해당 실험을 찾을 수 없습니다.")
    return obj


@router.patch("/experiments/{experiment_id}", response_model=schemas.ExperimentOut, summary="실험 수정")
def update_experiment(experiment_id: int, payload: schemas.ExperimentUpdate, db: Session = Depends(get_db)):
    try:
        obj = crud.update_experiment(db, experiment_id, payload)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    if not obj:
        raise HTTPException(status_code=404, detail="수정할 실험을 찾을 수 없습니다.")
    return obj


@router.delete("/experiments/{experiment_id}", summary="실험 삭제")
def delete_experiment(experiment_id: int, db: Session = Depends(get_db)):
    ok = crud.delete_experiment(db, experiment_id)
    if not ok:
        raise HTTPException(status_code=404, detail="삭제할 실험을 찾을 수 없습니다.")
    return {"ok": True}


# --- Result Schema API ---

@router.get("/projects/{project_id}/result-schemas", response_model=list[schemas.ResultSchemaOut], summary="프로젝트별 결과 스키마 조회")
def list_result_schemas(project_id: int, db: Session = Depends(get_db)):
    return crud.list_result_schemas(db, project_id)


@router.post("/result-schemas", response_model=schemas.ResultSchemaOut, summary="결과 스키마 생성")
def create_result_schema(payload: schemas.ResultSchemaCreate, db: Session = Depends(get_db)):
    proj = crud.get_project(db, payload.project_id)
    if not proj:
        raise HTTPException(status_code=400, detail="유효하지 않은 프로젝트 ID입니다.")
    return crud.create_result_schema(db, payload)


@router.patch("/result-schemas/{schema_id}", response_model=schemas.ResultSchemaOut, summary="결과 스키마 수정")
def update_result_schema(schema_id: int, payload: schemas.ResultSchemaUpdate, db: Session = Depends(get_db)):
    if payload.value_type == "categorical" and not payload.options:
        raise HTTPException(status_code=422, detail="categorical 타입은 options가 필요합니다.")

    obj = crud.update_result_schema(db, schema_id, payload)
    if not obj:
        raise HTTPException(status_code=404, detail="해당 결과 스키마를 찾을 수 없습니다.")
    return obj


@router.delete("/result-schemas/{schema_id}", summary="결과 스키마 삭제")
def delete_result_schema(schema_id: int, db: Session = Depends(get_db)):
    ok = crud.delete_result_schema(db, schema_id)
    if not ok:
        raise HTTPException(status_code=404, detail="삭제할 결과 스키마를 찾을 수 없습니다.")
    return {"ok": True}


# --- Output config API ---

@router.get("/projects/{project_id}/output-config", response_model=schemas.OutputConfigOut | None, summary="프로젝트별 출력 설정 조회")
def get_output_config(project_id: int, db: Session = Depends(get_db)):
    return crud.get_output_config(db, project_id)


@router.put("/output-config", response_model=schemas.OutputConfigOut, summary="출력 설정 저장")
def upsert_output_config(payload: schemas.OutputConfigUpsert, db: Session = Depends(get_db)):
    proj = crud.get_project(db, payload.project_id)
    if not proj:
        raise HTTPException(status_code=400, detail="유효하지 않은 프로젝트 ID입니다.")
    return crud.upsert_output_config(db, payload)


# --- Ingredient API ---

@router.get("/ingredients", response_model=list[schemas.IngredientOut], summary="원료 목록/검색")
def list_ingredients(
    q: str | None = Query(default=None),
    limit: int = Query(default=50, ge=1, le=200),
    db: Session = Depends(get_db),
):
    return crud.list_ingredients(db, q=q, limit=limit)


# --- CSV/XLSX IO helpers ---

def _to_rows_from_csv(raw: bytes) -> list[dict[str, Any]]:
    text = raw.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for row in reader:
        rows.append({k: (v if v is not None else "") for k, v in row.items()})
    return rows


def _to_rows_from_xlsx(raw: bytes) -> list[dict[str, Any]]:
    workbook = load_workbook(io.BytesIO(raw), data_only=True)
    sheet = workbook.active
    values = list(sheet.values)
    if not values:
        return []

    headers = [str(v).strip() if v is not None else "" for v in values[0]]
    rows: list[dict[str, Any]] = []
    for row in values[1:]:
        record: dict[str, Any] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            record[header] = row[idx] if idx < len(row) else ""
        if any(str(v).strip() for v in record.values() if v is not None):
            rows.append(record)
    return rows


def _to_csv_bytes(rows: list[dict[str, Any]]) -> bytes:
    out = io.StringIO()
    writer = csv.DictWriter(out, fieldnames=crud.EXPORT_COLUMNS)
    writer.writeheader()
    for row in rows:
        writer.writerow({key: row.get(key, "") for key in crud.EXPORT_COLUMNS})
    return out.getvalue().encode("utf-8-sig")


def _to_xlsx_bytes(rows: list[dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "data"
    ws.append(crud.EXPORT_COLUMNS)
    for row in rows:
        ws.append([row.get(key, "") for key in crud.EXPORT_COLUMNS])

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# --- CSV/XLSX import/export API ---

@router.get("/data/template", summary="업로드 템플릿 다운로드")
def download_template(format: Literal["csv", "xlsx"] = Query(default="csv")):
    sample = {
        "project_name": "샘플 프로젝트",
        "project_type": "REGULAR",
        "project_status": "ONGOING",
        "expected_end_date": "",
        "experiment_name": "실험-001",
        "author": "홍길동",
        "purpose": "조건 탐색",
        "experiment_conditions": "25C, 30분 교반",
        "experiment_date": "2026-03-10",
        "requester": "내부",
        "received_date": "2026-03-09",
        "materials_json": '[{"name":"원료A","amount":10,"unit":"g","ratio":50},{"name":"원료B","amount":10,"unit":"g","ratio":50}]',
        "result_values_json": '{"viscosity":1200,"status":"ok"}',
    }
    rows = [sample]

    if format == "csv":
        payload = _to_csv_bytes(rows)
        return StreamingResponse(
            io.BytesIO(payload),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=experiment_import_template.csv"},
        )

    payload = _to_xlsx_bytes(rows)
    return StreamingResponse(
        io.BytesIO(payload),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=experiment_import_template.xlsx"},
    )


@router.get("/data/export", summary="프로젝트/실험 데이터 다운로드")
def export_data(
    format: Literal["csv", "xlsx"] = Query(default="csv"),
    project_id: int | None = Query(default=None),
    db: Session = Depends(get_db),
):
    rows = crud.export_flat_rows(db, project_id=project_id)

    if format == "csv":
        payload = _to_csv_bytes(rows)
        filename = "eln_export.csv" if project_id is None else f"eln_project_{project_id}.csv"
        return StreamingResponse(
            io.BytesIO(payload),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    payload = _to_xlsx_bytes(rows)
    filename = "eln_export.xlsx" if project_id is None else f"eln_project_{project_id}.xlsx"
    return StreamingResponse(
        io.BytesIO(payload),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.post("/data/import", response_model=schemas.DataImportSummary, summary="CSV/XLSX 업로드로 데이터 재적용/업데이트")
async def import_data(
    format: Literal["csv", "xlsx"] = Query(default="csv"),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    raw = await file.read()
    if not raw:
        raise HTTPException(status_code=400, detail="파일이 비어 있습니다.")

    try:
        rows = _to_rows_from_csv(raw) if format == "csv" else _to_rows_from_xlsx(raw)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"파일 파싱에 실패했습니다: {exc}") from exc

    if not rows:
        raise HTTPException(status_code=400, detail="유효한 데이터 행이 없습니다.")

    try:
        return crud.import_flat_rows(db, rows)
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"데이터 반영에 실패했습니다: {exc}") from exc

